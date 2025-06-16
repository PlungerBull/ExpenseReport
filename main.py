import pandas as pd
import os
import shutil
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.table import Table
import win32com.client
import time
import ctypes

################### PATH MANAGER ###################
JSON_FILE_PATH = "C:/Users/Public/paths.json"

# Read the JSON file and load its content into the 'paths' dictionary
try:
    with open(JSON_FILE_PATH, 'r', encoding='utf-8') as f:
        paths = json.load(f)
except FileNotFoundError:
    print(f"Error: The JSON file was not found at '{JSON_FILE_PATH}'. Please check the path.")
    exit()
except json.JSONDecodeError:
    print(f"Error: Could not decode JSON from '{JSON_FILE_PATH}'. Please check the file's content.")
    exit()
except Exception as e:
    print(f"An unexpected error occurred while loading the JSON file: {e}")
    exit()

######################################################################

def process_expense_reports(original_file_path, template_file_path, output_directory):
    """
    Reads an Excel file, filters data by 'owner', creates new Excel files,
    pastes data and formulas, converts the range to a an Excel Table.
    Also filters the 'cecoOwner' sheet based on the current owner.
    """

    report_period = input("Please enter the period for the report (e.g., 2023_Q4): ")
    if not report_period:
        print("Report period cannot be empty. Exiting.")
        return

    # Define the columns to read from the ORIGINAL Expense Report Main.xlsx:
    # We read ALL columns from A to P from the source file because the 'owner'
    # column (which is a formula column in the template) is used for filtering.
    # The source file *must* contain 'owner' data for the filter to work.
    cols_to_read_indices_original = list(range(16)) # This covers A:P (indices 0-15)

    print(f"Reading 'dataRaw' sheet (columns A-P) from '{original_file_path}'...")
    try:
        df_data_raw_original = pd.read_excel(original_file_path, sheet_name='dataRaw', usecols=cols_to_read_indices_original)
        print("Successfully read 'dataRaw' sheet (columns A-P).")
    except KeyError:
        print("Error: 'dataRaw' sheet not found in the Excel file. Please ensure the sheet name is 'dataRaw'.")
        return
    except Exception as e:
        print(f"An error occurred while reading the original Excel file: {e}")
        return

    # The filtering for 'owner' still relies on the 'owner' column being present.
    # This 'owner' column in df_data_raw_original should contain actual values from the source file.
    if "owner" not in df_data_raw_original.columns:
        print("Error: 'owner' column not found in 'dataRaw' sheet. Please ensure 'owner' is the header for column N in your source data.")
        return

    unique_owners = df_data_raw_original['owner'].dropna().unique().tolist()
    if not unique_owners:
        print("No unique owners found in the 'owner' column. Exiting.")
        return
    print(f"Found {len(unique_owners)} unique owners.")
    print(f"The unique owners are: {unique_owners}")

    # --- Pre-read formulas from the template only ONCE ---
    try:
        template_wb_temp = openpyxl.load_workbook(template_file_path, data_only=False)
        template_ws_temp = template_wb_temp['dataContable']

        formulas_to_copy = {}
        formula_start_row = 2  # Formulas are assumed to be in row 2 of the template
        # Formulas are now in M, N, O, P (indices 12, 13, 14, 15)
        # M: helper, N: owner, O: subOwner, P: centroCostoDescrip
        formula_cols = [13, 14, 15, 16] # M, N, O, P (1-based index)

        for col_idx in formula_cols:
            cell = template_ws_temp.cell(row=formula_start_row, column=col_idx)
            if cell.data_type == 'f':  # 'f' means formula
                formulas_to_copy[col_idx] = cell.value
            else:
                # If a formula is expected but not found, this warning will trigger.
                # This is a good sanity check for your template.
                print(
                    f"Warning: Cell {get_column_letter(col_idx)}{formula_start_row} in template is not a formula. It will not be copied down.")
                formulas_to_copy[col_idx] = None  # Mark as None if no formula

        template_wb_temp.close()
        print("Successfully pre-read formulas from template.")
    except KeyError:
        print(
            f"Error: 'dataContable' sheet not found in the template file '{template_file_path}'. Please check sheet name.")
        return
    except Exception as e:
        print(f"An error occurred while pre-reading formulas from template: {e}")
        return

    # Ensure output directory exists
    os.makedirs(output_directory, exist_ok=True)

    for owner in unique_owners:
        clean_owner_name = "".join([c for c in str(owner) if c.isalnum() or c in (' ', '-', '_')]).replace(' ', '_')
        new_file_name = f"ExpenseReport_{clean_owner_name}_{report_period}.xlsx"
        new_file_path = os.path.join(output_directory, new_file_name)

        print(f"\nProcessing data for owner: '{owner}' and creating file: '{new_file_name}'...")

        df_filtered_for_owner_all_cols = df_data_raw_original[df_data_raw_original['owner'] == owner].copy()

        # Columns to write back are A to L (indices 0 to 11)
        # 'lineP&L' (L, index 11) is the last data column.
        cols_to_write_back = df_filtered_for_owner_all_cols.columns[:12].tolist()
        df_filtered_for_owner = df_filtered_for_owner_all_cols[cols_to_write_back]

        try:
            # 1. Copy the template file to the new destination
            shutil.copy2(template_file_path, new_file_path)
            print(f"Copied template to '{new_file_name}'.")

            # 2. Load the copied workbook
            wb = openpyxl.load_workbook(new_file_path)
            ws_data = wb['dataContable']

            # --- Processing 'cecoOwner' sheet ---
            try:
                ws_ceco_owner = wb['cecoOwner']
                print(f"Processing 'cecoOwner' sheet for owner: '{owner}'...")

                # Read the headers from the first row
                headers_ceco = [cell.value for cell in ws_ceco_owner[1]]

                # Read the data from the sheet starting from row 2
                data_ceco = []
                for r_idx in range(2, ws_ceco_owner.max_row + 1):
                    row_values = [ws_ceco_owner.cell(row=r_idx, column=c_idx).value for c_idx in
                                  range(1, ws_ceco_owner.max_column + 1)]
                    data_ceco.append(row_values)

                df_ceco = pd.DataFrame(data_ceco, columns=headers_ceco)

                if not df_ceco.empty:
                    # Filter out rows where 'mainOwner' is not equal to the current 'owner'
                    if 'mainOwner' in df_ceco.columns:
                        df_ceco_filtered_data = df_ceco[
                            df_ceco['mainOwner'].astype(str).str.strip().str.lower() == str(owner).strip().lower()
                            ].copy()
                    else:
                        print("Warning: 'mainOwner' column not found in 'cecoOwner' sheet. Skipping filtering.")
                        df_ceco_filtered_data = pd.DataFrame(columns=headers_ceco)  # Create empty df with headers

                    # Clear existing data rows (from row 2 onwards)
                    print("Clearing existing data rows in 'cecoOwner' sheet from row 2 onwards...")
                    for r_idx in range(ws_ceco_owner.max_row, 1, -1):  # From max_row down to 2
                        ws_ceco_owner.delete_rows(r_idx)
                    print("Finished clearing 'cecoOwner' data rows.")

                    # Write filtered data rows back to 'cecoOwner' sheet, starting from row 2
                    print(f"Writing {len(df_ceco_filtered_data)} filtered rows to 'cecoOwner' sheet, starting row 2.")
                    for r_idx, row_data in enumerate(df_ceco_filtered_data.values.tolist()):
                        for c_idx, cell_value in enumerate(row_data):
                            ws_ceco_owner.cell(row=r_idx + 2, column=c_idx + 1,
                                               value=cell_value)  # +2 because we start from row 2
                    print("Finished writing filtered data to 'cecoOwner'.")
                else:
                    print("Warning: 'cecoOwner' sheet is empty or only has headers. No filtering applied.")

            except KeyError:
                print(f"Warning: 'cecoOwner' sheet not found in '{new_file_name}'. Skipping filtering for this sheet.")
            except Exception as e:
                print(f"An error occurred while processing 'cecoOwner' sheet: {e}")

            # 3. Clear existing data (A-L) from row 2 onwards on dataContable sheet
            print(f"Clearing existing data in columns A-L from row 2 onwards...")
            # Now clearing 12 columns (A to L, 1-based index 1 to 12)
            for r_idx in range(2, ws_data.max_row + 1):
                for c_idx in range(1, 13): # Columns A to L (1 to 12)
                    cell_to_clear = ws_data.cell(row=r_idx, column=c_idx)
                    if cell_to_clear.value is not None:
                        cell_to_clear.value = None
            print("Finished clearing data.")

            # 4. Write headers (A-L) to the first row (A1 to L1) on dataContable sheet
            headers = list(df_filtered_for_owner.columns)
            for col_idx, header_name in enumerate(headers):
                ws_data.cell(row=1, column=col_idx + 1, value=header_name)

            # 5. Write filtered data (A-L) starting from row 2 on dataContable sheet
            start_row_for_data = 2
            min_col_data = 1

            data_to_write = df_filtered_for_owner.values.tolist()
            num_rows_to_write = len(data_to_write)

            print(f"Writing {num_rows_to_write} rows of data to columns A-L, starting from row {start_row_for_data}...")
            for r_idx, row_data in enumerate(data_to_write):
                if not row_data and len(row_data) == 0:
                    continue

                for c_idx, cell_value in enumerate(row_data):
                    if c_idx < len(df_filtered_for_owner.columns):
                        ws_data.cell(row=start_row_for_data + r_idx, column=min_col_data + c_idx, value=cell_value)
                    else:
                        break
            print("Finished writing data.")

            # 6. Copy Formulas Down in Columns M, N, O, P on dataContable sheet
            last_data_row = start_row_for_data + num_rows_to_write - 1
            if num_rows_to_write == 0:
                last_data_row = 1

            print(f"Copying formulas down to row {last_data_row} for columns M-P...")
            for col_idx_formula in formula_cols:
                original_formula = formulas_to_copy.get(col_idx_formula)
                if original_formula:
                    source_cell_coords = f"{get_column_letter(col_idx_formula)}{formula_start_row}"

                    for target_row in range(formula_start_row, last_data_row + 1):
                        target_cell_coords = f"{get_column_letter(col_idx_formula)}{target_row}"
                        translated_formula = Translator(original_formula, source_cell_coords).translate_formula(
                            target_cell_coords)
                        ws_data.cell(row=target_row, column=col_idx_formula, value=translated_formula)
                else:
                    pass # This 'pass' is important: if formula_to_copy[col_idx] was None due to the warning, it won't try to copy.
            print("Finished copying formulas.")

            # 7. Convert range to Excel Table "rawData" on dataContable sheet
            table_name = "rawData"
            # last_column_of_table_ref_index is now column P (index 15)
            last_column_of_table_ref_index = max(formula_cols) # This will be 16 for 'P'

            if num_rows_to_write == 0:
                table_end_row = 1  # Table will only include header row
            else:
                table_end_row = last_data_row

            new_table_ref = f"A1:{get_column_letter(last_column_of_table_ref_index)}{table_end_row}"

            tables_to_remove = []
            for table_entry in ws_data._tables:
                # openpyxl returns Table objects or strings for tables
                if isinstance(table_entry, Table) and table_entry.name == table_name:
                    tables_to_remove.append(table_entry)
                elif isinstance(table_entry, str) and table_entry == table_name:
                    # In older openpyxl or certain scenarios, it might be just the name string
                    tables_to_remove.append(table_entry)
            for item in tables_to_remove:
                ws_data._tables.remove(item)
                print(f"Removed existing table definition '{table_name}' from internal list before re-creating.")

            tab = Table(displayName=table_name, ref=new_table_ref)
            ws_data.add_table(tab)
            print(f"Created Excel Table '{table_name}' with range '{new_table_ref}'.")

            # 8. Go to the 'ExpenseReport' sheet and write 'pivotTableSource' in cell C2
            try:
                ws_pivot = wb['ExpenseReport']
                ws_pivot['C2'] = "rawData"
                print("Written 'pivotTableSource' to cell C2 of 'ExpenseReport' sheet.")
            except KeyError:
                print(f"Warning: 'ExpenseReport' sheet not found in the workbook. Could not write to C2.")
            except Exception as e:
                print(f"An error occurred while writing to 'ExpenseReport' sheet: {e}")

            # 9. Eliminate the 'dummySheet'
            if 'dummySheet' in wb.sheetnames:
                del wb['dummySheet']
                print("Removed 'dummySheet' from the workbook.")
            else:
                print("Warning: 'dummySheet' not found in the workbook. No sheet was removed.")

            # 10. Save the modified workbook
            wb.save(new_file_path)
            print(f"Successfully updated file for owner '{owner}' and saved to '{new_file_name}'.")

        except KeyError as ke:
            print(f"Error: Missing sheet in the template file. Check sheet names ('dataContable'): {ke}")
        except Exception as e:
            print(f"An unexpected error occurred while processing file for owner '{owner}' at '{new_file_path}': {e}")

    print("\nAll expense reports generated successfully!")
    print(f"You can find the generated files in: {output_directory}")

def calculate_total_saldo_soles(output_directory, template_file_path):
    """
    Calculates the total 'Saldo Soles' from all generated Excel files in the
    output directory, excluding the template file.

    Args:
        output_directory (str): The path to the folder where processed files are.
        template_file_path (str): The full path to the template file.

    Returns:
        float: The sum of 'Saldo Soles' from all relevant files.
    """
    total_saldo_soles = 0.0
    template_file_name = os.path.basename(template_file_path)

    print(f"\n--- Calculating total 'Saldo Soles' from files in: {output_directory} ---")
    excel_files = [f for f in os.listdir(output_directory) if f.endswith(('.xlsx', '.xlsm'))]

    if not excel_files:
        print("No Excel files found in the output folder to calculate sum.")
        return 0.0

    for file_name in excel_files:
        if file_name == template_file_name:
            print(f"  Skipping template file: '{file_name}'")
            continue

        file_path = os.path.join(output_directory, file_name)
        try:
            # Read 'Saldo Soles' column from 'dataContable' sheet
            # Saldo Soles is column K (index 10)
            df_temp = pd.read_excel(file_path, sheet_name='dataContable', usecols=["Saldo Soles"])
            if 'Saldo Soles' in df_temp.columns:
                current_file_sum = df_temp['Saldo Soles'].sum()
                total_saldo_soles += current_file_sum
                print(f"  Added {current_file_sum:.2f} from '{file_name}'.")
            else:
                print(f"  Warning: 'Saldo Soles' column not found in the newly created file. Skipping sum for this file.")
        except KeyError:
            print(f"  Error: 'dataContable' sheet not found in '{file_name}'. Skipping sum for this file.")
        except Exception as e:
            print(f"  Error reading 'Saldo Soles' from '{file_name}': {e}")

    print(f"--- Finished calculating total 'Saldo Soles'. Total: {total_saldo_soles:,.2f} ---")
    return total_saldo_soles


def refresh_excel_files_in_folder(folder_path):
    """
    Opens all Excel files in the specified folder, performs a 'Refresh All',
    saves, and closes them using COM automation.
    """
    if not os.path.isdir(folder_path):
        print(f"Error (Folder Refresh): Folder not found at '{folder_path}'. Skipping refresh.")
        return

    print(f"\n--- Starting 'Refresh All' for Excel files in: {folder_path} ---")
    excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xlsm'))]

    if not excel_files:
        print("No Excel files found in the folder to refresh.")
        return

    excel_app = None
    try:
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = True  # Keep Excel hidden
        excel_app.DisplayAlerts = False  # Suppress alerts (e.g., about external links)

        for file_name in excel_files:
            file_path = os.path.join(folder_path, file_name)
            workbook = None
            # Skip the template file during this refresh process as it's not a generated report.
            if file_name == os.path.basename(paths['templateExpenseReport']):
                print(f"  Skipping template file '{file_name}' from refresh cycle.")
                continue

            try:
                print(f"  Opening and refreshing: {file_name}")
                workbook = excel_app.Workbooks.Open(file_path)

                # --- ONLY REFRESH, NO PIVOTTABLE SOURCE CHANGE ---
                # The assumption here is that your template's PivotTable
                # is already correctly configured to refresh itself based on the
                # data in the 'dataContable' sheet, perhaps using a dynamic named range
                # or an existing table reference that updates automatically.

                workbook.RefreshAll()  # The core refresh command
                excel_app.CalculateUntilAsyncQueriesDone()  # Wait for queries to complete
                time.sleep(1) # Small pause after refresh

                workbook.Save()
                print(f"  Successfully refreshed and saved: {file_name}")
                time.sleep(3)
            except Exception as e:
                print(f"  Error processing '{file_name}': {e}")
            finally:
                if workbook:
                    workbook.Close(SaveChanges=False)

    except Exception as e:
        print(f"An error occurred with Excel COM application: {e}")
    finally:
        if excel_app:
            excel_app.Quit()
        print(f"--- Finished 'Refresh All' for files in: {folder_path} ---")


def move_files_to_history(output_directory, history_directory, template_file_path):
    """
    Moves all Excel files from the output_directory to the history_directory,
    excluding the template file if it exists in the output directory.

    Args:
        output_directory (str): The path to the folder where processed files are.
        history_directory (str): The path to the folder where files should be moved.
        template_file_path (str): The full path to the template file.
    """

    # Get just the template file name for exclusion
    template_file_name = os.path.basename(template_file_path)

    moved_count = 0
    print(f"\n--- Starting to move files to history: {history_directory} ---")
    for file_name in os.listdir(output_directory):
        if file_name.endswith(('.xlsx', '.xlsm')) and file_name != template_file_name:
            source_path = os.path.join(output_directory, file_name)
            destination_path = os.path.join(history_directory, file_name)
            try:
                shutil.move(source_path, destination_path)
                print(f"  Moved '{file_name}' to history.")
                moved_count += 1
            except Exception as e:
                print(f"  Error moving '{file_name}': {e}")

    if moved_count == 0:
        print("No generated Excel files found to move to history.")
    else:
        print(f"Successfully moved {moved_count} files to history.")
    print("--- Finished moving files ---")

def is_hidden_windows(filepath):
    """
    Checks if a file is hidden on Windows using ctypes to access file attributes.
    """
    if os.name == 'nt':
        try:
            FILE_ATTRIBUTE_HIDDEN = 0x02
            attrs = ctypes.windll.kernel32.GetFileAttributesW(str(filepath))
            return (attrs != -1) and (attrs & FILE_ATTRIBUTE_HIDDEN)
        except Exception as e:
            print(f"Warning: Could not check hidden attribute for '{filepath}': {e}")
            return False
    return False

# --- Helper Function to simulate Power Query's "Transform File" ---
def transform_excel_file(file_path):
    """
    Reads an Excel file (assuming first sheet) and returns a DataFrame.
    This simulates the Power Query 'Transform File' custom function.
    """
    try:
        df = pd.read_excel(file_path)
        return df
    except Exception as e:
        print(f"Error reading Excel file '{file_path}': {e}")
        return pd.DataFrame()

### Main Data Processing Function
def process_sales_reports(input_folder_path, output_folder_path, period):
    """
    Consolidates and transforms data from Excel files to generate a sales report
    and saves the result to an Excel file.

    Args:
        input_folder_path (str): The path to the folder containing source Excel files.
        output_folder_path (str): The path to the folder where the output Excel file will be saved.
        period (str): The period string (e.g., '2023-12') used for naming the output file.
    """

    output_filename = f"{period}. SalesReport.xlsx"
    output_full_path = os.path.join(output_folder_path, output_filename)

    all_data_frames = []

    print(f"Scanning for Excel files in: {input_folder_path}")
    for entry in os.scandir(input_folder_path):
        if entry.is_file() and entry.name.endswith('.xlsx') and not is_hidden_windows(entry.path):
            file_path = entry.path
            file_name = entry.name
            print(f"Processing file: {file_name}")

            transformed_df = transform_excel_file(file_path)

            if not transformed_df.empty:
                # Extract company name from file name
                name_parts = file_name.replace('.xlsx', '').split('-')
                company_name = name_parts[-1] if len(name_parts) > 2 else "UNKNOWN"

                transformed_df['company'] = company_name
                all_data_frames.append(transformed_df)

    if not all_data_frames:
        print("No valid Excel files found or processed in the input directory.")
        return

    df_combined = pd.concat(all_data_frames, ignore_index=True)

    # Filter out 'FLN' company
    # df_combined = df_combined[df_combined['company'] != "FLN"].copy()

    # Change column types
    type_mapping = {
        "Fecha": 'datetime64[ns]',
        "Débito Local": float,
        "Débito Dólar": float,
        "Crédito Local": float,
        "Crédito Dolar": float
    }
    for col, dtype in type_mapping.items():
        if col in df_combined.columns:
            df_combined[col] = pd.to_numeric(df_combined[col], errors='coerce')
            if dtype == 'datetime64[ns]':
                df_combined[col] = pd.to_datetime(df_combined[col], errors='coerce')
            else:
                 df_combined[col] = df_combined[col].astype(dtype)


    # Filter rows where 'Cuenta Contable' starts with "70"
    if 'Cuenta Contable' in df_combined.columns:
        df_combined = df_combined[df_combined['Cuenta Contable'].astype(str).str.startswith("70", na=False)].copy()

    # Add 'saldoPEN' column
    df_combined['Crédito Local'] = df_combined['Crédito Local'].fillna(0)
    df_combined['Débito Local'] = df_combined['Débito Local'].fillna(0)
    df_combined['saldoPEN'] = df_combined['Crédito Local'] - df_combined['Débito Local']

    # Clean 'Fuente' column
    if 'Fuente' in df_combined.columns:
        df_combined['Fuente'] = df_combined['Fuente'].astype(str)
        df_combined['Fuente'] = (
            df_combined['Fuente']
            .str.replace("INTERFASCE ODOO ", "", regex=False)
            .str.replace("INTERFACE ODOO ", "", regex=False)
            .str.replace("N/C", "", regex=False)
            .str.replace("FAC", "", regex=False)
            .str.replace("B/V", "", regex=False)
            .str.replace("#", "", regex=False)
            .str.replace(" ", "", regex=False)
        )

    # Remove specified columns
    columns_to_remove = [
        "Tipo De Documento", "Documento", "Débito Local", "Débito Dólar",
        "Crédito Local", "Crédito Dolar", "Centro Costo", "Consecutivo",
        "Tipo De Asiento", "Glosa", "Descripción Tipo De Asiento", "Módulo",
        "Usuario Impresión", "Flujo Efectivo", "Patrimonio Neto", "Proyecto",
        "Descripción de Proyecto", "Fase", "Descripción de Fase"
    ]
    df_final = df_combined.drop(columns=[col for col in columns_to_remove if col in df_combined.columns], errors='ignore')

    # Save to Excel
    try:
        df_final.to_excel(output_full_path, index=False, sheet_name='rawData')
        print(f"\nSuccessfully processed data and saved to: {output_full_path}")

        os.startfile(output_full_path)  # This line will open the Excel file

    except Exception as e:
        print(f"Error saving or opening the Excel file: {e}")

# --- Main execution block ---
if __name__ == "__main__":
    original_excel_file_path = paths['ExpenseReport']
    template_excel_file_path = paths['templateExpenseReport']
    output_folder = paths['outputExpenseReport']
    history_folder = paths['expenseReportHistory']
    sales_report_path = paths['salesDataStorage']
    expense_report_data_actual_path = paths['expenseReportDataActual']

    #### EXPENSE REPORT ####
    # move_files_to_history(output_folder, history_folder, template_excel_file_path)
    # process_expense_reports(original_excel_file_path, template_excel_file_path, output_folder)
    # refresh_excel_files_in_folder(output_folder)
    # final_total_expense_soles = calculate_total_saldo_soles(output_folder, template_excel_file_path)
    # print(f"\nFINAL TOTAL EXPENSE FOR THE PERIOD: {final_total_expense_soles:,.2f}")

    #### SALES REPORT ####
    period_input = str(input("Please enter the period (e.g., '2023-12'): "))
    process_sales_reports(expense_report_data_actual_path, sales_report_path, period_input)
    move_files_to_history(sales_report_path, history_folder)