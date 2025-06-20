import pandas as pd
import os
import shutil
import openpyxl
import win32com.client
import time
import pyodbc

def _apply_company_replacements(df: pd.DataFrame, replacements_map: dict, df_name: str) -> pd.DataFrame:
    """Applies company name replacements to a DataFrame."""
    if replacements_map:
        print(f"Applying company name replacements to {df_name} data...")
        for old_value, new_value in replacements_map.items():
            initial_count = (df['company'] == old_value).sum()
            if initial_count > 0:
                df['company'] = df['company'].replace(old_value, new_value)
                final_count = (df['company'] == old_value).sum()
                print(f"  Replaced '{old_value}' with '{new_value}'. {initial_count - final_count} occurrences replaced in {df_name} data.")
            else:
                print(f"  '{old_value}' not found in 'company' column of {df_name} data. No replacement made for this value.")
    return df

def _clear_excel_range(sheet, start_row: int, max_col_idx: int, clear_until_row: int):
    """Clears a specified range in an Openpyxl worksheet."""
    for row_num in range(start_row, clear_until_row + 1):
        for col_num in range(3, max_col_idx + 3): # Assuming data starts at column C (index 3)
            sheet.cell(row=row_num, column=col_num).value = None

def template_forecast_generator(access_db_path: str, excel_template_path: str, output_folder_path: str):
    """
    Accesses tables in an Access database, filters and transforms data,
    separates content by unique 'subOwner' values, pastes each subset into
    a template Excel file, and saves it with a dynamic name.
    This version includes processing for both 'forecastExpenses' and 'forecastHeadcount' sheets.
    For 'forecastExpenses', it extends formulas in columns AH and AI down.
    For 'forecastHeadcount', it pivots headcount data and pastes it into the designated sheet,
    now directly filtering by 'subOwner' and 'mainOwner' from the headcount data itself.

    This optimized version consolidates database connections, pre-loads all data,
    and uses helper functions to reduce code duplication.

    Args:
        access_db_path (str): The full path to the Access database file.
        excel_template_path (str): The full path to the Excel template file.
        output_folder_path (str): The path to the folder where output Excel files will be saved.
    """

    # 1. Prompt for User Input
    user_forecast_version = input("Please enter the forecasting version (e.g., '6+6', 'Q3'): ")
    if not user_forecast_version:
        print("Forecasting version cannot be empty. Exiting.")
        return

    # Static parameters
    access_expense_table_name = "fullDetailedP&L"
    access_headcount_table_name = "headcountFull"
    column_for_separation = "subOwner" # This remains the primary column for file separation
    excel_expenses_sheet_name = "forecastExpenses"
    excel_headcount_sheet_name = "forecastHeadcount"
    template_header_row = 7 # Not directly used in data pasting, but good for context
    data_start_row = 8
    fixed_identifier_columns_expenses = [
        "company", "lineP&L", "centroCosto",
        "description", "cuentaContable", "descriptionCuentaContable"
    ]
    fixed_identifier_columns_headcount = [
        "company", "centroCosto", "description", "jobGeneral"
    ]

    initial_columns_from_db_expenses = fixed_identifier_columns_expenses + ["mainOwner", "subOwner", "periodo", "saldoPEN"]
    initial_columns_from_db_headcount = ["company", "period", "nameID", "centroCosto", "jobGeneral", "description", "mainOwner", "subOwner"]

    company_replacements_map = {
        'ROP': 'FLXTECH'
    }

    if not all([access_db_path, excel_template_path, output_folder_path]):
        print("Error: One or more required paths passed as arguments are missing or None.")
        print(f"Access DB Path: {access_db_path}")
        print(f"Excel Template Path: {excel_template_path}")
        print(f"Output Folder Path: {output_folder_path}")
        return

    print("\n--- Starting Data Processing ---")
    print(f"Access Database: {access_db_path}")
    print(f"Expense Table Name: {access_expense_table_name}")
    print(f"Headcount Table Name: {access_headcount_table_name}")
    print(f"Splitting by: {column_for_separation}")
    print(f"Template Expense Sheet: {excel_expenses_sheet_name}")
    print(f"Template Headcount Sheet: {excel_headcount_sheet_name}")

    os.makedirs(output_folder_path, exist_ok=True)

    df_expenses = pd.DataFrame()
    df_headcount = pd.DataFrame()
    cnxn = None # Initialize connection object
    excel_app = None # Initialize Excel application object for win32com

    try:
        # --- Consolidated Database Connection and Data Loading ---
        print("\n--- Connecting to Access database and retrieving all data ---")
        conn_str = (
            r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
            f"DBQ={access_db_path};"
        )
        cnxn = pyodbc.connect(conn_str)
        print("Database connection established.")

        # Load Expense Data
        columns_to_select_str_expenses = ", ".join(f"[{col}]" for col in initial_columns_from_db_expenses)
        query_expenses = f"SELECT {columns_to_select_str_expenses} FROM [{access_expense_table_name}]"
        df_expenses = pd.read_sql(query_expenses, cnxn)
        print("Expense data retrieved.")

        # Load Headcount Data
        columns_to_select_str_headcount = ", ".join(f"[{col}]" for col in initial_columns_from_db_headcount)
        query_headcount = f"SELECT {columns_to_select_str_headcount} FROM [{access_headcount_table_name}]"
        df_headcount = pd.read_sql(query_headcount, cnxn)
        print("Headcount data retrieved.")

        cnxn.close()
        print("Database connection closed.")

        # --- Apply company replacements to both datasets using helper function ---
        df_expenses = _apply_company_replacements(df_expenses, company_replacements_map, "expense")
        df_headcount = _apply_company_replacements(df_headcount, company_replacements_map, "headcount")

        # --- Expense Data Transformations (Now applied to the full df_expenses) ---
        initial_rows_expenses = len(df_expenses)
        df_expenses = df_expenses[~df_expenses['cuentaContable'].astype(str).str.startswith('62')].copy()
        rows_after_filter_expenses = len(df_expenses)
        print(f"Filtered 'cuentaContable' for expenses: Removed {initial_rows_expenses - rows_after_filter_expenses} rows where 'cuentaContable' started with '62'.")

        print("Performing expense data transformations...")
        df_expenses['periodo'] = pd.to_datetime(df_expenses['periodo'])
        df_expenses['periodo'] = df_expenses['periodo'].dt.to_period('M').dt.end_time
        df_expenses['periodo_month_name'] = df_expenses['periodo'].dt.strftime('%B')

        id_vars_for_pivot_expenses = fixed_identifier_columns_expenses + ["mainOwner", "subOwner"]
        df_unpivoted_expenses_full = df_expenses.pivot_table(
            index=id_vars_for_pivot_expenses,
            columns='periodo_month_name',
            values='saldoPEN',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        month_names_ordered = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]

        # --- Headcount Data Transformations (Now applied to the full df_headcount) ---
        print("Performing headcount data transformations on full dataset...")
        df_headcount['period'] = pd.to_datetime(df_headcount['period'])
        df_headcount['period'] = df_headcount['period'].dt.to_period('M').dt.end_time
        df_headcount['period_month_name'] = df_headcount['period'].dt.strftime('%B')

        headcount_pivot_index = fixed_identifier_columns_headcount + ["mainOwner", "subOwner"] # Include for full pivot
        df_pivoted_headcount_full = df_headcount.pivot_table(
            index=headcount_pivot_index,
            columns='period_month_name',
            values='nameID',
            aggfunc='count',
            fill_value=0
        ).reset_index()


        print("Data pre-processing complete. Ready to generate Excel files.")

        unique_subowners = df_unpivoted_expenses_full[column_for_separation].unique()

        if len(unique_subowners) == 0:
            print("No unique 'subOwner' values found. No Excel files will be generated.")
            return

        print(f"Found {len(unique_subowners)} unique {column_for_separation} values. Processing each...")

        try:
            # Get Excel Application object (only once)
            try:
                excel_app = win32com.client.GetActiveObject("Excel.Application")
            except:
                excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            print("Excel application instance obtained for win32com operations.")

            for i, subowner in enumerate(unique_subowners):
                print(f"\nProcessing '{subowner}' ({i + 1}/{len(unique_subowners)})...")

                # --- Determine file paths and copy template ---
                sanitized_subowner = "".join(c for c in str(subowner) if c.isalnum() or c in [' ', '_', '-']).strip()
                output_filename = f"Forecast_{user_forecast_version}_{sanitized_subowner}.xlsx"
                output_filepath = os.path.join(output_folder_path, output_filename)

                try:
                    shutil.copy(excel_template_path, output_filepath)
                    print(f"Copied template to: {output_filepath}")
                except Exception as e:
                    print(f"Error copying template for '{subowner}': {e}")
                    continue

                # --- Openpyxl for pasting data (both sheets) ---
                try:
                    workbook = openpyxl.load_workbook(output_filepath)

                    # --- Expense Sheet Processing with openpyxl ---
                    print(f"Processing '{excel_expenses_sheet_name}' sheet...")
                    df_filtered_expenses = df_unpivoted_expenses_full[df_unpivoted_expenses_full[column_for_separation] == subowner].copy()

                    unique_main_owners_for_subowner = df_filtered_expenses['mainOwner'].unique()

                    if len(unique_main_owners_for_subowner) > 1:
                        main_owner_for_cell_D4 = ", ".join(map(str, unique_main_owners_for_subowner))
                        print(f"Warning: Multiple mainOwners found for subOwner '{subowner}'. Using: {main_owner_for_cell_D4}")
                    elif len(unique_main_owners_for_subowner) == 1:
                        main_owner_for_cell_D4 = str(unique_main_owners_for_subowner[0])
                    else:
                        main_owner_for_cell_D4 = "N/A"

                    sub_owner_for_cell_D5 = str(subowner)

                    existing_month_columns_expenses = [col for col in month_names_ordered if col in df_filtered_expenses.columns]
                    final_data_table_columns_order_expenses = fixed_identifier_columns_expenses + existing_month_columns_expenses
                    df_filtered_for_paste_expenses = df_filtered_expenses[final_data_table_columns_order_expenses]
                    num_data_rows_expenses = len(df_filtered_for_paste_expenses)

                    sheet_expenses = workbook[excel_expenses_sheet_name]

                    sheet_expenses['D4'].value = main_owner_for_cell_D4
                    sheet_expenses['D5'].value = sub_owner_for_cell_D5
                    print(f"Main Owner ('{main_owner_for_cell_D4}') and Sub Owner ('{sub_owner_for_cell_D5}') written to D4/D5 on '{excel_expenses_sheet_name}'.")

                    data_to_paste_expenses = df_filtered_for_paste_expenses.values.tolist()

                    # Clear previous data in expense data table range using helper function
                    _clear_excel_range(sheet_expenses, data_start_row, len(final_data_table_columns_order_expenses), data_start_row + 1000)

                    # Write expense data rows
                    for r_idx, row_data in enumerate(data_to_paste_expenses):
                        for c_idx, value in enumerate(row_data):
                            sheet_expenses.cell(row=data_start_row + r_idx, column=c_idx + 3, value=value)

                    print(f"Main data table for '{excel_expenses_sheet_name}' pasted.")


                    # --- Headcount Data Processing and Pasting with openpyxl ---
                    print(f"\nProcessing '{excel_headcount_sheet_name}' sheet for '{subowner}'...")
                    df_pivoted_filtered_headcount = df_pivoted_headcount_full[
                        (df_pivoted_headcount_full['subOwner'] == subowner)
                    ].copy()

                    if df_pivoted_filtered_headcount.empty:
                        print(f"No headcount data found for subOwner '{subowner}'. Skipping headcount sheet update.")
                    else:
                        existing_month_columns_headcount = [col for col in month_names_ordered if col in df_pivoted_filtered_headcount.columns]
                        final_data_table_columns_order_headcount = fixed_identifier_columns_headcount + existing_month_columns_headcount
                        df_pivoted_for_paste_headcount = df_pivoted_filtered_headcount[final_data_table_columns_order_headcount]
                        num_data_rows_headcount = len(df_pivoted_for_paste_headcount)

                        sheet_headcount = workbook[excel_headcount_sheet_name]
                        print(f"Opened sheet: '{excel_headcount_sheet_name}'.")

                        data_to_paste_headcount = df_pivoted_for_paste_headcount.values.tolist()

                        # Clear previous data in headcount data table range using helper function
                        _clear_excel_range(sheet_headcount, data_start_row, len(final_data_table_columns_order_headcount), data_start_row + 1000)

                        # Write headcount data rows
                        for r_idx, row_data in enumerate(data_to_paste_headcount):
                            for c_idx, value in enumerate(row_data):
                                sheet_headcount.cell(row=data_start_row + r_idx, column=c_idx + 3, value=value)

                        print(f"Headcount data table pasted on '{excel_headcount_sheet_name}'.")

                    # Save the workbook once after all openpyxl operations on both sheets are complete
                    workbook.save(output_filepath)
                    print(f"Workbook saved (Openpyxl).")

                except Exception as e:
                    print(f"Error during openpyxl processing for '{subowner}': {e}")
                    continue

                # 8. Excel Manipulation with win32com.client (for refreshing and formula copying)
                try:
                    workbook_com = excel_app.Workbooks.Open(output_filepath)
                    sheet_com_expenses = workbook_com.Sheets(excel_expenses_sheet_name)

                    if num_data_rows_expenses > 0:
                        print("Extending formulas in columns AH and AI of 'forecastExpenses' using win32com...")
                        source_range = sheet_com_expenses.Range("AH8:AI8")
                        last_formula_row_expenses = data_start_row + num_data_rows_expenses - 1
                        if last_formula_row_expenses >= data_start_row + 1:
                            destination_range_str_expenses = f"AH{data_start_row + 1}:AI{last_formula_row_expenses}"
                            destination_range_expenses = sheet_com_expenses.Range(destination_range_str_expenses)
                            source_range.Copy()
                            destination_range_expenses.PasteSpecial(Paste=-4123) # xlPasteFormulas
                            excel_app.CutCopyMode = False # Clear clipboard
                            print(f"Formulas from AH8:AI8 copied down to {destination_range_str_expenses} on 'forecastExpenses'.")
                        else:
                            print("Not enough expense data rows to extend formulas on 'forecastExpenses'.")
                    else:
                        print("No expense data rows pasted, skipping formula extension on 'forecastExpenses'.")

                    print(f"Refreshing all formulas and connections in '{output_filename}' (win32com)...")
                    workbook_com.RefreshAll()

                    time.sleep(5) # Give Excel some time to refresh

                    workbook_com.Save()
                    workbook_com.Close(False)
                    print(f"Workbook refreshed and saved (win32com).")

                except Exception as e:
                    print(f"Error during win32com Excel automation for '{subowner}': {e}")

        finally: # This finally ensures excel_app is quit at the very end of the function's execution
            if excel_app is not None:
                try:
                    if excel_app.Workbooks.Count == 0:
                        excel_app.Quit()
                        print("Excel application quit.")
                    else:
                        print("Excel application not quit (other workbooks are open).")
                except Exception as e:
                    print(f"Error quitting Excel application: {e}")

    except Exception as e:
        print(f"An error occurred during the main process: {e}")
    finally:
        if cnxn and not cnxn.closed: # Ensure connection is closed if an error occurred before explicit close
            cnxn.close()
            print("Database connection explicitly closed in final block.")
        print("\n--- Data Processing Complete ---")