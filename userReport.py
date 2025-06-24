import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# ################### PATH MANAGER ###################
JSON_FILE_PATH = "C:/Users/Public/paths.json"  # Adjust this path as needed

# Read the JSON file and load its content into the 'paths' dictionary.
# This section will crash if the file is not found, JSON is malformed, or key is missing.
with open(JSON_FILE_PATH, 'r', encoding='utf-8') as f:
    paths = json.load(f)



# ################### END PATH MANAGER ###################

def _write_alerts_section(ws, df, main_header_text, start_col_index):
    """
    Helper function to write a section to the ALERTS worksheet.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The ALERTS worksheet object.
        df (pd.DataFrame): The DataFrame containing data for this section.
        main_header_text (str): The text for the main bolded header.
        start_col_index (int): The 1-based starting column index for this section.
    """
    # Main header starts at ROW 2
    ws.cell(row=2, column=start_col_index).value = main_header_text
    ws.cell(row=2, column=start_col_index).font = Font(bold=True)

    if not df.empty:
        # Secondary headers start at ROW 4 and are now BOLDED
        for col_offset, col_name in enumerate(df.columns):
            cell = ws.cell(row=4, column=start_col_index + col_offset)
            cell.value = col_name
            cell.font = Font(bold=True)  # Make secondary headers bold

        # Data starts at ROW 5
        for r_offset, row_data in enumerate(dataframe_to_rows(df, index=False, header=False)):
            for c_offset, cell_value in enumerate(row_data):
                ws.cell(row=5 + r_offset, column=start_col_index + c_offset).value = cell_value


def process_and_alert_client_data(report_dir):
    """
    Processes an Excel file:
    1. Renames the initial sheet to 'rawData'.
    2. Creates a 'Clients by status' sheet with filtered data.
    3. Creates an 'ALERTS' sheet with dependent filtered data and unique values.
    4. Creates a 'Summary' sheet based on 'Clients by status' data.
    5. Ensures 'ALERTS' is first, 'Summary' is second, and 'Clients by status' is third.
    All modifications are done within a single workbook session.
    """
    excel_file_path = None
    try:
        files_in_directory = [f for f in os.listdir(report_dir) if os.path.isfile(os.path.join(report_dir, f))]
        for file_name in files_in_directory:
            if file_name.endswith(('.xlsx', '.xls')):
                excel_file_path = os.path.join(report_dir, file_name)
                break
        if not excel_file_path:
            print(f"Error: No Excel file found in {report_dir}")
            return
    except FileNotFoundError:
        print(f"Error: Directory not found: {report_dir}")
        return
    except Exception as e:
        print(f"Error finding Excel file: {e}")
        return

    # --- SINGLE WORKBOOK SESSION STARTS HERE ---
    try:
        wb = load_workbook(excel_file_path)

        # Get a reference to the initial first sheet. This is our 'rawData' source.
        initial_source_sheet_name = wb.sheetnames[0]
        ws_initial_source = wb[initial_source_sheet_name]

        # Read data directly from this initial sheet into df_original
        data_rows = ws_initial_source.values
        headers = next(data_rows)
        df_original = pd.DataFrame(list(data_rows), columns=headers)
        print(f"Data loaded from '{initial_source_sheet_name}' sheet in-memory successfully.")

        # ################### START Empresa Value Replacements ###################
        empresa_replacements = {
            "FIBERLUX TECH SOCIEDAD ANONIMA CERRADA": "FLXTECH",
            "NEXTNET S.A.C.": "NXT",
            "FIBERLUX SOCIEDAD ANONIMA CERRADA": "FLX"
        }

        if "Empresa" in df_original.columns:
            # Apply the replacements to the 'Empresa' column
            df_original["Empresa"] = df_original["Empresa"].replace(empresa_replacements)
            print("Empresa column values replaced successfully.")
        else:
            print("Warning: 'Empresa' column not found in df_original for replacement.")

        # 1. Create the new target sheets first. This ensures the workbook always has sheets.
        ws_alerts = wb.create_sheet("ALERTS")
        ws_summary = wb.create_sheet("Summary")
        ws_clients = wb.create_sheet("Clients by status")
        print("New target sheets created.")

        # Collect all names of sheets that we want to remove from the original workbook.
        sheets_to_remove_if_old = [initial_source_sheet_name, "ALERTS", "Summary", "Clients by status"]

        # Iterate over a *copy* of sheetnames, as we are modifying the collection while iterating
        for sheet_name in list(wb.sheetnames):
            if sheet_name in sheets_to_remove_if_old:
                # IMPORTANT: Ensure we are NOT deleting the *newly created* sheets.
                if wb[sheet_name] not in [ws_alerts, ws_summary, ws_clients]:
                    del wb[sheet_name]
                    print(f"Old sheet '{sheet_name}' deleted from workbook (in memory).")

    except Exception as e:
        print(f"An error occurred during initial Excel file handling (load or read): {e}")
        return

    # Define all required columns for validation and filtering
    all_potential_columns = [
        "Sale Subscription Line ID", "Linea", "Servicio", "Tipo Servicio",
        "Estado Servicio", "Moneda", "Venta Solarizada", "ID Servicio Contrato",
        "Contrato", "Plantilla Contrato", "Cliente", "Equipoventa",
        "Empresa", "Cliente Ruc Dni"
    ]
    for col in all_potential_columns:
        if col not in df_original.columns:
            print(f"Warning: Expected column '{col}' not found in the original Excel file.")

    # --- Prepare "Clients by status" Data (Main Sheet) ---
    client_status_cols = [
        "Sale Subscription Line ID", "Linea", "Servicio", "Tipo Servicio",
        "Estado Servicio", "Moneda", "Venta Solarizada", "ID Servicio Contrato",
        "Contrato", "Plantilla Contrato", "Cliente", "Equipoventa",
        "Empresa", "Cliente Ruc Dni"
    ]
    client_status_actual_cols = [col for col in client_status_cols if col in df_original.columns]

    df_clients_by_status = df_original[client_status_actual_cols].copy()

    if "Tipo Servicio" in df_clients_by_status.columns:
        df_clients_by_status = df_clients_by_status[df_clients_by_status["Tipo Servicio"] == "Recurrente"]
    else:
        print("Warning: 'Tipo Servicio' column not found for 'Clients by status' filter.")

    if "Plantilla Contrato" in df_clients_by_status.columns:
        df_clients_by_status = df_clients_by_status[~df_clients_by_status["Plantilla Contrato"].isin([
            "PAGO EN CUOTAS-NEXTNET", "PAGO EN CUOTAS-TECH y NEXTNET"
        ])]
    else:
        print("Warning: 'Plantilla Contrato' column not found for 'Clients by status' filter.")

    # --- Prepare "ALERTS" Sheet Sections (Dependent Filters) ---

    # Define common exclusion list for "Estado Servicio"
    excluded_estados = ["Baja", "Baja Adm"]

    # NEW Filtro 1 Data: Clients with Plantilla Contrato = PAGOS EN CUOTA
    filtro1_new_cols = ["Empresa", "Estado Servicio", "Cliente", "Equipoventa", "Plantilla Contrato"]
    filtro1_new_actual_cols = [col for col in filtro1_new_cols if col in df_original.columns]

    df_filtro1_new = pd.DataFrame()
    clientes_in_filtro1_new = pd.Series(dtype='object')
    if "Plantilla Contrato" in df_original.columns:
        df_temp_filtro1 = df_original[df_original["Plantilla Contrato"].isin([
            "PAGO EN CUOTAS-NEXTNET", "PAGO EN CUOTAS-TECH y NEXTNET"
        ])]
        # ################### MODIFIED: Exclude "Baja" and "Baja Adm" from Filtro 1 ###################
        if "Estado Servicio" in df_temp_filtro1.columns:
            df_temp_filtro1 = df_temp_filtro1[~df_temp_filtro1["Estado Servicio"].isin(excluded_estados)]
        else:
            print("Warning: 'Estado Servicio' column not found in df_temp_filtro1 for exclusion.")

        df_filtro1_new = df_temp_filtro1[filtro1_new_actual_cols].drop_duplicates(subset=["Cliente"]).copy()
        if "Cliente" in df_filtro1_new.columns:
            clientes_in_filtro1_new = df_filtro1_new["Cliente"]
    else:
        print("Warning: 'Plantilla Contrato' column not found for 'Filtro 1' (new) data.")

    # NEW Filtro 2 Data: Clients with Tipo Servicio <> Recurrente AND NOT in NEW Filtro 1
    filtro2_new_cols = ["Empresa", "Estado Servicio", "Cliente", "Equipoventa", "Tipo Servicio"]
    filtro2_new_actual_cols = [col for col in filtro2_new_cols if col in df_original.columns]

    df_filtro2_new = pd.DataFrame()
    if "Tipo Servicio" in df_original.columns:
        df_temp_filtro2 = df_original[df_original["Tipo Servicio"] != "Recurrente"]

        if "Cliente" in df_temp_filtro2.columns and not clientes_in_filtro1_new.empty:
            df_temp_filtro2 = df_temp_filtro2[~df_temp_filtro2["Cliente"].isin(clientes_in_filtro1_new)]

        # ################### MODIFIED: Exclude "Baja" and "Baja Adm" from Filtro 2 ###################
        if "Estado Servicio" in df_temp_filtro2.columns:
            df_temp_filtro2 = df_temp_filtro2[~df_temp_filtro2["Estado Servicio"].isin(excluded_estados)]
        else:
            print("Warning: 'Estado Servicio' column not found in df_temp_filtro2 for exclusion.")
        # ################### END MODIFIED ###################

        df_filtro2_new = df_temp_filtro2[filtro2_new_actual_cols].drop_duplicates(subset=["Cliente"]).copy()
    else:
        print("Warning: 'Tipo Servicio' column not found for 'Filtro 2' (new) data.")

    # Unique Plantilla Contrato Data: From the 'Clients by status' df
    df_plantilla_contrato_unique = pd.DataFrame()
    if "Plantilla Contrato" in df_clients_by_status.columns:
        unique_plantilla_values = df_clients_by_status["Plantilla Contrato"].dropna().unique()
        df_plantilla_contrato_unique = pd.DataFrame(unique_plantilla_values,
                                                    columns=["Plantilla Contrato (Unique Values)"])
    else:
        print("Warning: 'Plantilla Contrato' column not found for 'Unique Plantilla Contrato' data.")

    # --- Prepare "Summary" Sheet Data (with Empresa and Equipoventa on rows, Empresa sorted) ---
    df_summary = pd.DataFrame()
    required_summary_cols = ["Empresa", "Equipoventa", "Estado Servicio", "Cliente"]

    if all(col in df_clients_by_status.columns for col in required_summary_cols):
        # Create a copy to avoid SettingWithCopyWarning when setting categorical dtypes
        df_summary_base = df_clients_by_status[required_summary_cols].copy()

        # Define custom order for Empresa (rows)
        # IMPORTANT: Use the NEW (abbreviated) values here for sorting and display
        desired_empresa_order = ["FLXTECH", "NXT", "FLX"]
        all_unique_empresas = df_summary_base['Empresa'].unique().tolist()
        final_empresas_sort_order = [emp for emp in desired_empresa_order if emp in all_unique_empresas] + \
                                    sorted([emp for emp in all_unique_empresas if emp not in desired_empresa_order])

        # Apply categorical dtype for Empresa to ensure custom sort order
        df_summary_base['Empresa'] = pd.Categorical(df_summary_base['Empresa'],
                                                    categories=final_empresas_sort_order,
                                                    ordered=True)

        # Define custom order for Estado Servicio (columns)
        desired_estado_servicio_order = ["Activo", "Suspendido", "Instalacion", "Baja", "Baja Adm"]
        all_unique_estados = df_summary_base['Estado Servicio'].unique().tolist()
        final_estado_servicio_cols_order = [est for est in desired_estado_servicio_order if est in all_unique_estados] + \
                                           sorted([est for est in all_unique_estados if
                                                   est not in desired_estado_servicio_order])

        # Apply categorical dtype for Estado Servicio to ensure custom sort order
        df_summary_base['Estado Servicio'] = pd.Categorical(df_summary_base['Estado Servicio'],
                                                            categories=final_estado_servicio_cols_order,
                                                            ordered=True)

        # Group by Empresa, Equipoventa, and Estado Servicio, then count distinct Clients
        df_summary = df_summary_base.groupby(["Empresa", "Equipoventa", "Estado Servicio"])[
            "Cliente"].nunique().unstack(fill_value=0)

        # Sort the index explicitly by the categorical 'Empresa' and then by 'Equipoventa'
        df_summary = df_summary.sort_index(level=['Empresa', 'Equipoventa'])

        # Ensure all desired Estado Servicio columns are present (even if empty) and ordered
        df_summary = df_summary.reindex(columns=final_estado_servicio_cols_order, fill_value=0)

        # --- NEW STEP: Omit rows where all Estado Servicio counts are zero ---
        # Identify the columns that contain the counts (Estado Servicio types)
        count_columns_for_sum = [col for col in final_estado_servicio_cols_order if col in df_summary.columns]

        # Sum across these columns for each row
        df_summary_non_zero = df_summary[df_summary[count_columns_for_sum].sum(axis=1) > 0].copy()

        # Calculate Grand Total Row based on the filtered data
        grand_total_row = df_summary_non_zero.sum().to_frame().T
        grand_total_row.index = pd.MultiIndex.from_tuples([("Grand Total", "")], names=["Empresa", "Equipoventa"])

        # Concatenate Grand Total to the filtered DataFrame
        df_summary = pd.concat([df_summary_non_zero, grand_total_row])

        # Reset index to make 'Empresa' and 'Equipoventa' regular columns for writing
        df_summary = df_summary.reset_index()
        # Rename the index columns for clarity in Excel. '' for Equipoventa keeps it blank as shown.
        df_summary.rename(columns={'Empresa': 'Row Labels', 'Equipoventa': ''}, inplace=True)

    else:
        print(
            "Warning: Missing required columns ('Empresa', 'Equipoventa', 'Estado Servicio', or 'Cliente') for 'Summary' data.")

    # --- Final Excel Sheet Writing and Saving ---
    try:
        # Write data for "Clients by status" sheet, bolding row 1
        rows_to_write_clients = dataframe_to_rows(df_clients_by_status, index=False, header=True)

        header_row_clients = next(rows_to_write_clients)
        ws_clients.append(header_row_clients)
        for cell in ws_clients[1]:
            cell.font = Font(bold=True)

        for row_data in rows_to_write_clients:
            ws_clients.append(row_data)

        # Write data for "ALERTS" sheet using the helper function
        _write_alerts_section(ws_alerts, df_filtro1_new,
                              "FILTRO 1: CLIENTES CON PLANTILLA CONTRATO = PAGOS EN CUOTA", 1)  # Column A (1)

        _write_alerts_section(ws_alerts, df_filtro2_new,
                              "FILTRO 2: CLIENTES CON TIPO SERVICIO <> RECURRENTE (NO EXCLUIDOS POR FILTRO 1)",
                              11)  # Column K (11)

        _write_alerts_section(ws_alerts, df_plantilla_contrato_unique,
                              "UNIQUE VALUES FOR PLANTILLA CONTRATO IN THE ORIGINAL DATA", 21)  # Column U (21)

        # Write data for "Summary" sheet
        # "SUMMARY:" on row 2, column B (2)
        ws_summary.cell(row=2, column=2).value = "SUMMARY:"
        ws_summary.cell(row=2, column=2).font = Font(bold=True)

        if not df_summary.empty:
            # Write headers for summary starting at Row 4, Column B (2)
            current_cols_for_header = df_summary.columns.tolist()
            for col_idx, col_name in enumerate(current_cols_for_header, 2):  # Start from column B
                ws_summary.cell(row=4, column=col_idx).value = col_name
                ws_summary.cell(row=4, column=col_idx).font = Font(bold=True)  # Make summary headers bold

            # --- MODIFIED: Write data for summary starting at Row 5, Column B (2) ---
            last_row_label = None  # To track the previous 'Empresa' value (which is 'Row Labels' column)
            start_data_excel_row = 5  # Data starts from Excel row 5
            excel_col_offset = 2  # Data starts at Excel Column B (index 2)

            # Iterate through DataFrame rows directly to control cell writing for 'Row Labels'
            for r_offset, row_tuple in enumerate(df_summary.itertuples(index=False)):
                current_row_label = row_tuple[0]  # 'Row Labels' is the first element

                # Determine value for 'Row Labels' column (Column B in Excel)
                cell_value_to_write = ""
                # Write the label if it's the Grand Total or if it's a new group (not repeating)
                if current_row_label == "Grand Total" or current_row_label != last_row_label:
                    cell_value_to_write = current_row_label
                last_row_label = current_row_label

                # Write the 'Row Labels' column
                cell_row_label = ws_summary.cell(row=start_data_excel_row + r_offset, column=excel_col_offset)
                cell_row_label.value = cell_value_to_write

                # Write the rest of the columns (starting from the second column in df_summary, which is Excel Column C)
                # row_tuple[1:] skips the 'Row Labels' column
                for c_offset, cell_value in enumerate(row_tuple[1:]):
                    # Column C in Excel is index 3. So, excel_col_offset + 1 + c_offset.
                    cell = ws_summary.cell(row=start_data_excel_row + r_offset, column=excel_col_offset + 1 + c_offset)
                    cell.value = cell_value

                # Apply bold to the entire "Grand Total" row if it's the Grand Total row
                if current_row_label == "Grand Total":
                    for col_num in range(excel_col_offset, len(current_cols_for_header) + excel_col_offset):
                        ws_summary.cell(row=start_data_excel_row + r_offset, column=col_num).font = Font(bold=True)

        # Reorder sheets: ALERTS (1st), Summary (2nd), Clients by status (3rd), (remaining sheets if any)
        wb.move_sheet(ws_alerts, offset=-len(wb.sheetnames) + 1)
        wb.move_sheet(ws_summary, offset=-len(wb.sheetnames) + 2)
        wb.move_sheet(ws_clients, offset=-len(wb.sheetnames) + 3)

        # Save the workbook (final save)
        wb.save(excel_file_path)
        print(
            f"Excel file '{excel_file_path}' updated with 'ALERTS' (1st), 'Summary' (2nd), and 'Clients by status' (3rd) sheets.")

    except Exception as e:
        print(f"An error occurred during final Excel file manipulation: {e}")